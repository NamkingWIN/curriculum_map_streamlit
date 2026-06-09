import sys
import re
from collections import defaultdict, deque
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================
# CẤU HÌNH CHO DATABASE CHUẨN HÓA
# =========================

SHEET_COURSE = "course"
SHEET_RELATION_TYPE = "relation"
SHEET_COURSE_RELATION = "course_relation"

OUTPUT_FILE_NAME = "KET_QUA_KIEM_TRA.xlsx"

# Mặc định: nếu môn điều kiện ở cùng học kỳ với môn đang xét thì cần cảnh báo.
ALLOW_SAME_SEMESTER_HOC_TRUOC = False
ALLOW_SAME_SEMESTER_PHAI_DAT_TRUOC = False


# =========================
# HÀM TIỆN ÍCH
# =========================

def safe_str(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_code(value) -> str:
    """Chuẩn hóa mã học phần / mã quan hệ để tránh lỗi 11001001.0."""
    if pd.isna(value):
        return ""

    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()

    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]

    text = text.upper()
    text = re.sub(r"\s+", "", text)
    return text


def parse_semester(value):
    """Chuyển giá trị học kỳ sang số để so sánh."""
    if pd.isna(value):
        return None

    text = str(value).strip()
    if not text:
        return None

    match = re.search(r"\d+", text)
    if match:
        return int(match.group())

    return None


def is_truthy(value) -> bool:
    """Đọc các cột is_active/source_found... dạng 1/0, TRUE/FALSE, YES/NO."""
    if pd.isna(value):
        return False

    text = str(value).strip().upper()
    return text in ["1", "TRUE", "YES", "Y", "CO", "CÓ", "ACTIVE"]


def require_columns(df: pd.DataFrame, sheet_name: str, columns: list[str]):
    missing = [col for col in columns if col not in df.columns]
    if missing:
        raise ValueError(
            f"Sheet '{sheet_name}' thiếu cột bắt buộc: {', '.join(missing)}"
        )


def normalize_relation_type(value) -> str:
    """Chuẩn hóa tên loại quan hệ để tương thích draw_flow.py."""
    text = safe_str(value).lower()

    if "học trước" in text or "hoc truoc" in text:
        return "Học trước"

    if (
        "tiên quyết" in text
        or "tien quyet" in text
        or "phải đạt" in text
        or "phai dat" in text
    ):
        return "Phải đạt trước"

    return safe_str(value)


def relation_allow_same_semester(relation_type: str) -> bool:
    if relation_type == "Học trước":
        return ALLOW_SAME_SEMESTER_HOC_TRUOC
    if relation_type == "Phải đạt trước":
        return ALLOW_SAME_SEMESTER_PHAI_DAT_TRUOC
    return False


# =========================
# ĐỌC DATABASE
# =========================

def read_database(input_file: Path):
    """Đọc dữ liệu từ file database chuẩn hóa."""
    xls = pd.ExcelFile(input_file)
    required_sheets = [SHEET_COURSE, SHEET_RELATION_TYPE, SHEET_COURSE_RELATION]
    missing_sheets = [sheet for sheet in required_sheets if sheet not in xls.sheet_names]

    if missing_sheets:
        raise ValueError(
            "File không đúng cấu trúc database chuẩn hóa. "
            f"Thiếu sheet: {', '.join(missing_sheets)}"
        )

    course_df = pd.read_excel(input_file, sheet_name=SHEET_COURSE, dtype=object)
    relation_type_df = pd.read_excel(input_file, sheet_name=SHEET_RELATION_TYPE, dtype=object)
    course_relation_df = pd.read_excel(input_file, sheet_name=SHEET_COURSE_RELATION, dtype=object)

    require_columns(
        course_df,
        SHEET_COURSE,
        ["course_id", "course_name_vi", "semester"],
    )
    require_columns(
        relation_type_df,
        SHEET_RELATION_TYPE,
        ["relation_type_id", "relation_type_name"],
    )
    require_columns(
        course_relation_df,
        SHEET_COURSE_RELATION,
        ["target_course_id", "source_course_id", "relation_type_id"],
    )

    return course_df, relation_type_df, course_relation_df


def build_course_dictionary(course_df: pd.DataFrame) -> dict:
    """Tạo dictionary tra cứu học phần theo course_id."""
    courses = {}

    for index, row in course_df.iterrows():
        course_id = normalize_code(row.get("course_id"))
        if not course_id:
            continue

        is_active = row.get("is_active", 1)
        if "is_active" in course_df.columns and not is_truthy(is_active):
            continue

        semester_raw = row.get("semester")
        semester = parse_semester(semester_raw)

        if course_id in courses:
            courses[course_id]["Dòng trùng mã"] += f", sheet course dòng {index + 2}"
            continue

        courses[course_id] = {
            "Mã học phần": course_id,
            "Tên học phần": safe_str(row.get("course_name_vi")),
            "Số tín chỉ": row.get("credits", ""),
            "Học kỳ gốc": semester_raw,
            "Học kỳ": semester,
            "Loại học phần": safe_str(row.get("course_type")),
            "Ngoài CTĐT": row.get("outside_curriculum", ""),
            "Thứ tự hiển thị": row.get("display_order", ""),
            "Dòng Excel": f"sheet course dòng {index + 2}",
            "Dòng trùng mã": "",
            "Ghi chú": safe_str(row.get("note")),
        }

    return courses


def build_relation_type_lookup(relation_type_df: pd.DataFrame) -> dict:
    """Tạo dictionary tra cứu loại quan hệ."""
    relation_types = {}

    for _, row in relation_type_df.iterrows():
        relation_type_id = normalize_code(row.get("relation_type_id"))
        if not relation_type_id:
            continue

        relation_types[relation_type_id] = {
            "relation_type_id": relation_type_id,
            "relation_type_name": normalize_relation_type(row.get("relation_type_name")),
            "source_column": safe_str(row.get("source_column")),
            "description": safe_str(row.get("description")),
        }

    return relation_types


# =========================
# KIỂM TRA QUAN HỆ
# =========================

def check_relations(course_relation_df: pd.DataFrame, courses: dict, relation_types: dict):
    """Kiểm tra mã học phần điều kiện, thứ tự học kỳ và tạo bảng quan hệ đầy đủ."""
    results = []
    all_relations = []
    missing_codes = []

    for index, row in course_relation_df.iterrows():
        if "is_active" in course_relation_df.columns and not is_truthy(row.get("is_active", 1)):
            continue

        relation_id = safe_str(row.get("relation_id")) or f"REL_ROW_{index + 2}"
        target = normalize_code(row.get("target_course_id"))
        source = normalize_code(row.get("source_course_id"))
        relation_type_id = normalize_code(row.get("relation_type_id"))
        relation_type = relation_types.get(relation_type_id, {}).get("relation_type_name", relation_type_id)
        source_column = safe_str(row.get("source_column")) or relation_types.get(relation_type_id, {}).get("source_column", "")

        target_course = courses.get(target)
        source_course = courses.get(source)

        relation = {
            "Dòng Excel": f"sheet course_relation dòng {index + 2}",
            "Mã quan hệ": relation_id,
            "Mã học phần": target,
            "Tên học phần": target_course["Tên học phần"] if target_course else "",
            "Học kỳ học phần": target_course["Học kỳ"] if target_course else "",
            "Học kỳ gốc học phần": target_course["Học kỳ gốc"] if target_course else "",
            "Mã học phần điều kiện": source,
            "Tên học phần điều kiện": source_course["Tên học phần"] if source_course else "",
            "Học kỳ học phần điều kiện": source_course["Học kỳ"] if source_course else "",
            "Học kỳ gốc học phần điều kiện": source_course["Học kỳ gốc"] if source_course else "",
            "Dòng Excel học phần điều kiện": source_course["Dòng Excel"] if source_course else "",
            "Loại điều kiện": relation_type,
            "Mã loại điều kiện": relation_type_id,
            "Cột nguồn": source_column,
        }

        missing_parts = []
        if not target_course:
            missing_parts.append(f"mã học phần đang xét {target}")
        if not source_course:
            missing_parts.append(f"mã học phần điều kiện {source}")

        if missing_parts:
            relation.update({
                "Mức độ": "Mã không tìm thấy",
                "Nhận xét": (
                    "Không tìm thấy " + ", ".join(missing_parts) +
                    " trong sheet course. Cần kiểm tra lại bảng course_relation hoặc danh mục course."
                ),
            })
            results.append(relation)
            missing_codes.append(relation)
            all_relations.append(relation)
            continue

        target_semester = target_course["Học kỳ"]
        source_semester = source_course["Học kỳ"]

        if target_semester is None or source_semester is None:
            relation.update({
                "Mức độ": "Không đủ dữ liệu học kỳ",
                "Nhận xét": (
                    "Một trong hai học phần không có dữ liệu học kỳ hợp lệ, "
                    "nên chưa thể so sánh."
                ),
            })
            results.append(relation)

        elif source_semester > target_semester:
            relation.update({
                "Mức độ": "Mâu thuẫn nặng",
                "Nhận xét": (
                    f"Học phần điều kiện {source} ở học kỳ {source_semester}, "
                    f"nhưng học phần {target} lại ở học kỳ {target_semester}. "
                    "Môn điều kiện đang được xếp sau môn cần học."
                ),
            })
            results.append(relation)

        elif source_semester == target_semester:
            if relation_allow_same_semester(relation_type):
                relation.update({
                    "Mức độ": "Không mâu thuẫn",
                    "Nhận xét": "Hai học phần cùng học kỳ, nhưng cấu hình trong mã nguồn đang cho phép trường hợp này.",
                })
            else:
                relation.update({
                    "Mức độ": "Cùng học kỳ - cần kiểm tra",
                    "Nhận xét": (
                        f"Học phần điều kiện {source} và học phần {target} "
                        f"cùng ở học kỳ {target_semester}. Nếu yêu cầu là học trước/đạt trước "
                        "thì cần xem lại."
                    ),
                })
                results.append(relation)

        else:
            relation.update({
                "Mức độ": "Không mâu thuẫn",
                "Nhận xét": "Học phần điều kiện được xếp ở học kỳ trước.",
            })

        all_relations.append(relation)

    redundant_conditions = find_redundant_conditions(all_relations, courses)
    results.extend(redundant_conditions)

    return results, all_relations, missing_codes, redundant_conditions


# =========================
# KIỂM TRA ĐIỀU KIỆN CÓ THỂ BỊ DƯ
# =========================

def build_graph_from_relations(all_relations: list[dict]) -> dict[str, set[str]]:
    graph = defaultdict(set)

    for relation in all_relations:
        if relation.get("Mức độ") == "Mã không tìm thấy":
            continue

        source = normalize_code(relation.get("Mã học phần điều kiện"))
        target = normalize_code(relation.get("Mã học phần"))

        if source and target and source != target:
            graph[source].add(target)

    return graph


def find_alternative_path(graph: dict[str, set[str]], source: str, target: str):
    """
    Tìm đường đi gián tiếp source -> ... -> target sau khi bỏ cạnh trực tiếp source -> target.
    Trả về một path nếu có, ngược lại trả về None.
    """
    queue = deque([[source]])
    visited = {source}

    while queue:
        path = queue.popleft()
        current = path[-1]

        for next_node in sorted(graph.get(current, set())):
            # Bỏ cạnh trực tiếp đang xét để tránh tự chứng minh chính nó.
            if current == source and next_node == target:
                continue

            if next_node in visited:
                continue

            new_path = path + [next_node]

            if next_node == target:
                if len(new_path) >= 3:
                    return new_path
                continue

            visited.add(next_node)
            queue.append(new_path)

    return None


def find_redundant_conditions(all_relations: list[dict], courses: dict) -> list[dict]:
    """
    Kiểm tra điều kiện có thể bị dư do quan hệ bắc cầu.

    Ví dụ:
        A -> B -> C
    thì quan hệ trực tiếp A -> C có thể bị dư vì chỉ cần giữ B -> C.
    """
    graph = build_graph_from_relations(all_relations)
    redundant = []
    seen = set()

    for relation in all_relations:
        if relation.get("Mức độ") == "Mã không tìm thấy":
            continue

        source = normalize_code(relation.get("Mã học phần điều kiện"))
        target = normalize_code(relation.get("Mã học phần"))

        if not source or not target or source == target:
            continue

        key = (source, target, relation.get("Loại điều kiện", ""), relation.get("Mã quan hệ", ""))
        if key in seen:
            continue
        seen.add(key)

        path = find_alternative_path(graph, source, target)
        if not path:
            continue

        direct_condition_to_keep = path[-2]
        chain_text = " → ".join(path)

        issue = dict(relation)
        issue.update({
            "Mức độ": "Điều kiện có thể bị dư",
            "Môn nên giữ trực tiếp": direct_condition_to_keep,
            "Tên môn nên giữ trực tiếp": courses.get(direct_condition_to_keep, {}).get("Tên học phần", ""),
            "Chuỗi quan hệ bao hàm": chain_text,
            "Nhận xét": (
                f"Môn {source} đang được khai báo là điều kiện trực tiếp của {target}. "
                f"Tuy nhiên đã có chuỗi quan hệ {chain_text}. "
                f"Vì vậy có thể chỉ cần giữ môn sau cùng {direct_condition_to_keep} "
                f"là điều kiện trực tiếp của {target}, còn {source} có thể bị dư trong danh sách điều kiện trực tiếp."
            ),
        })
        redundant.append(issue)

    return redundant


# =========================
# XUẤT KẾT QUẢ
# =========================

def make_output_file_name(input_file: Path) -> Path:
    return input_file.with_name(OUTPUT_FILE_NAME)


def style_excel(file_path: Path):
    wb = load_workbook(file_path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    error_fill = PatternFill("solid", fgColor="F4CCCC")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    missing_fill = PatternFill("solid", fgColor="D9EAD3")
    no_data_fill = PatternFill("solid", fgColor="D9D2E9")
    redundant_fill = PatternFill("solid", fgColor="FCE5CD")

    border = Border(
        left=Side(style="thin", color="D9EAD3"),
        right=Side(style="thin", color="D9EAD3"),
        top=Side(style="thin", color="D9EAD3"),
        bottom=Side(style="thin", color="D9EAD3"),
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers = [cell.value for cell in ws[1]]
        level_col = headers.index("Mức độ") + 1 if "Mức độ" in headers else None

        for row in range(2, ws.max_row + 1):
            fill = None

            if level_col:
                value = ws.cell(row=row, column=level_col).value

                if value == "Mâu thuẫn nặng":
                    fill = error_fill
                elif value == "Cùng học kỳ - cần kiểm tra":
                    fill = warning_fill
                elif value == "Mã không tìm thấy":
                    fill = missing_fill
                elif value == "Không đủ dữ liệu học kỳ":
                    fill = no_data_fill
                elif value == "Điều kiện có thể bị dư":
                    fill = redundant_fill

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                if fill:
                    cell.fill = fill

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0

            for cell in ws[col_letter]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 60)

    wb.save(file_path)


def write_output(
    input_file: Path,
    output_file: Path,
    courses: dict,
    results: list[dict],
    all_relations: list[dict],
    missing_codes: list[dict],
    redundant_conditions: list[dict],
):
    result_df = pd.DataFrame(results)
    all_relations_df = pd.DataFrame(all_relations)
    missing_df = pd.DataFrame(missing_codes)
    redundant_df = pd.DataFrame(redundant_conditions)
    course_df = pd.DataFrame(courses.values())

    total_courses = len(courses)
    total_relations = len(all_relations)
    total_issues = len(result_df)

    def count_level(level: str) -> int:
        if result_df.empty or "Mức độ" not in result_df.columns:
            return 0
        return int((result_df["Mức độ"] == level).sum())

    summary_df = pd.DataFrame([
        ["File kiểm tra", input_file.name],
        ["Định dạng dữ liệu", "Database chuẩn hóa"],
        ["Sheet học phần", SHEET_COURSE],
        ["Sheet loại quan hệ", SHEET_RELATION_TYPE],
        ["Sheet quan hệ học phần", SHEET_COURSE_RELATION],
        ["Tổng số học phần", total_courses],
        ["Tổng số quan hệ điều kiện", total_relations],
        ["Tổng số vấn đề cần xem lại", total_issues],
        ["Mâu thuẫn nặng", count_level("Mâu thuẫn nặng")],
        ["Cùng học kỳ - cần kiểm tra", count_level("Cùng học kỳ - cần kiểm tra")],
        ["Mã không tìm thấy", count_level("Mã không tìm thấy")],
        ["Không đủ dữ liệu học kỳ", count_level("Không đủ dữ liệu học kỳ")],
        ["Điều kiện có thể bị dư", count_level("Điều kiện có thể bị dư")],
    ], columns=["Nội dung", "Giá trị"])

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Tong hop", index=False)
        result_df.to_excel(writer, sheet_name="Ket qua kiem tra", index=False)
        redundant_df.to_excel(writer, sheet_name="Dieu kien co the bi du", index=False)
        missing_df.to_excel(writer, sheet_name="Ma khong tim thay", index=False)
        all_relations_df.to_excel(writer, sheet_name="Doi chieu tat ca", index=False)
        course_df.to_excel(writer, sheet_name="Danh muc hoc phan", index=False)

    style_excel(output_file)


# =========================
# MAIN
# =========================

def main():
    if len(sys.argv) < 2:
        print("Cách dùng:")
        print('python progress_check.py "progress_database_normalized.xlsx"')
        sys.exit(1)

    input_file = Path(sys.argv[1])

    if not input_file.exists():
        print(f"Không tìm thấy file: {input_file}")
        sys.exit(1)

    if input_file.suffix.lower() not in [".xlsx", ".xlsm"]:
        print("File đầu vào phải là .xlsx hoặc .xlsm")
        sys.exit(1)

    try:
        course_df, relation_type_df, course_relation_df = read_database(input_file)

        courses = build_course_dictionary(course_df)
        relation_types = build_relation_type_lookup(relation_type_df)

        results, all_relations, missing_codes, redundant_conditions = check_relations(
            course_relation_df=course_relation_df,
            courses=courses,
            relation_types=relation_types,
        )

        output_file = make_output_file_name(input_file)

        write_output(
            input_file=input_file,
            output_file=output_file,
            courses=courses,
            results=results,
            all_relations=all_relations,
            missing_codes=missing_codes,
            redundant_conditions=redundant_conditions,
        )

        print("Đã kiểm tra xong.")
        print(f"File kết quả: {output_file}")
        print("")
        print("Tóm tắt:")
        print("- Định dạng dữ liệu: Database chuẩn hóa")
        print(f"- Tổng số học phần: {len(courses)}")
        print(f"- Tổng số quan hệ điều kiện: {len(all_relations)}")
        print(f"- Tổng số vấn đề cần xem lại: {len(results)}")
        print(f"- Điều kiện có thể bị dư: {len(redundant_conditions)}")

    except Exception as e:
        print("Có lỗi khi xử lý:")
        print(str(e))
        sys.exit(1)


if __name__ == "__main__":
    main()
